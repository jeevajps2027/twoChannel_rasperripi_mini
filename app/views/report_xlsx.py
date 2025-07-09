from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from bs4 import BeautifulSoup
import os
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime
from pathlib import Path

def get_save_directory():
    usb_base_path = '/media'
    downloads_path = str(Path.home() / 'Downloads')

    # Check if any user folder exists inside /media
    if os.path.exists(usb_base_path):
        for user_folder in os.listdir(usb_base_path):
            user_path = os.path.join(usb_base_path, user_folder)
            if os.path.isdir(user_path):
                # Now check if any device inside user folder
                for device_folder in os.listdir(user_path):
                    device_path = os.path.join(user_path, device_folder)
                    if os.path.ismount(device_path):
                        return device_path, 'USB'

    # No USB found
    return downloads_path, 'Downloads'

@csrf_exempt
def report_xlsx(request):
    if request.method == 'POST':
        # Get form data
        from_date = request.POST.get('from_date', '')
        to_date = request.POST.get('to_date', '')
        mode = request.POST.get('mode', '')
        part_model = request.POST.get('part_model', '')
        shift = request.POST.get('shift', '')
        status = request.POST.get('status', '')
        total_count = request.POST.get('total_count', '')
        table_html = request.POST.get('table_html')

        if table_html:
            soup = BeautifulSoup(table_html, 'html.parser')
            thead = soup.find('thead')
            rows = soup.find_all('tr')

            wb = Workbook()
            ws = wb.active

            # Set headers
            ws['A1'] = f'From Date: {from_date}'
            ws['B1'] = f'To Date: {to_date}'
            ws['C1'] = f'Part Model: {part_model}'
            ws['D1'] = f'Mode: {mode}'
            ws['E1'] = f'Shift: {shift}'
            ws['F1'] = f'Total Count: {total_count}'
            ws['G1'] = f'Status: {status}'

            for cell in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1']:
                ws[cell].font = Font(bold=True)
                ws[cell].alignment = Alignment(horizontal='left')

            start_row = 5
            if thead:
                header_cells = thead.find_all('th')
                for col_num, th in enumerate(header_cells, start=1):
                    cell = ws.cell(row=start_row, column=col_num, value=th.text.strip())
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')

            for row_num, row in enumerate(rows[1:], start=start_row + 1):
                cells = row.find_all(['td', 'th'])
                for col_num, cell in enumerate(cells, start=1):
                    ws.cell(row=row_num, column=col_num, value=cell.text.strip())

            for col_num in range(1, ws.max_column + 1):
                max_length = 0
                col_letter = get_column_letter(col_num)
                for cell in ws[col_letter]:
                    try:
                        max_length = max(max_length, len(str(cell.value)) if cell.value else 0)
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2

            current_datetime = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            file_name = f'report_data_{current_datetime}.xlsx'

            # ✨ New Save Directory (USB or Downloads)
            save_dir, save_location = get_save_directory()
            os.makedirs(save_dir, exist_ok=True)
            file_path = os.path.join(save_dir, file_name)

            wb.save(file_path)

            return JsonResponse({'success': True, 'message': f'File saved to {save_location}: {file_path}'})
        else:
            return JsonResponse({'success': False, 'message': 'No table data provided.'})
    else:
        return JsonResponse({'success': False, 'message': 'Invalid request method.'})
